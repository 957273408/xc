import openpyxl
import json

wb = openpyxl.load_workbook(r'C:\Users\Administrator.DESKTOP-P8MU1HJ\Desktop\香肠\数据表格.xlsx', data_only=True)

result = []
for name in wb.sheetnames:
    ws = wb[name]
    header = [str(c.value) if c.value else '' for c in list(ws.rows)[0]]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(6, ws.max_row), values_only=True)):
        rows.append([str(v)[:100] if v is not None else '' for v in row])
    result.append({
        "name": name,
        "shape": [ws.max_row, ws.max_column],
        "header": header,
        "sample_rows": rows
    })
    print(f"Sheet: {name} | Rows: {ws.max_row} | Cols: {ws.max_column}")
    print(f"Header: {header}")
    for r in rows[:3]:
        print(f"  {r}")
    print()

with open(r'C:\Users\Administrator.DESKTOP-P8MU1HJ\Desktop\香肠\temp_xlsx_dump.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)
print("Done - saved to temp_xlsx_dump.json")
