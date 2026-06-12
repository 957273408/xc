import openpyxl
import json

file_path = r'c:\Users\Administrator.DESKTOP-P8MU1HJ\Desktop\香肠\战队信息表-2026-06-11 16_48_20(1).xlsx'
wb = openpyxl.load_workbook(file_path)
print('Sheets:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== Sheet: {sheet_name} ===')
    print(f'Dimensions: {ws.dimensions}')
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            print(row)
